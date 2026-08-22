#!/usr/bin/env python3
"""Build CapCut / Kling / Runway pricing comparison workbook (snapshot 2026-08-21)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = Path(__file__).with_name("video-ai-pricing-compare.xlsx")
AS_OF = "2026-08-21"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def style_header(ws, row: int, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 2, max_width))
        ws.column_dimensions[letter].width = max(width, 12)


def add_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list[str]]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for row in rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def main() -> None:
    wb = Workbook()
    meta = wb.active
    meta.title = "README"
    meta["A1"] = "Video AI pricing snapshot"
    meta["A1"].font = Font(bold=True, size=14)
    meta["A2"] = f"As of: {AS_OF} (UTC+9 research day)"
    meta["A3"] = "Services: CapCut, Kling AI, Runway"
    meta["A4"] = (
        "Caveat: prices vary by region, platform, tax, and promotions. "
        "Checkout / developer portal is authoritative."
    )
    meta["A4"].fill = NOTE_FILL
    meta["A5"] = (
        "Rule: app/web subscription credits are separate from API credits "
        "(Kling, Runway). CapCut has no public render/generation API rate card."
    )
    meta["A6"] = "Sheets: consumer, API, sources, commands"
    meta["A7"] = "Built by: docs/research/video-ai-pricing/build_pricing_xlsx.py"
    meta.column_dimensions["A"].width = 100

    add_sheet(
        wb,
        "consumer",
        ["service", "plan", "price_usd", "included", "notes"],
        [
            ["CapCut", "Free", "$0", "Basic editing", "Region/device dependent feature gates"],
            [
                "CapCut",
                "Standard",
                "~$9.99 / month",
                "Mid tier (often mobile-focused); less than Pro",
                "Listed on third-party roundups; confirm in-app checkout",
            ],
            [
                "CapCut",
                "Pro",
                "$19.99 / month or $179.99 / year",
                "Stronger AI toolkit, ~1TB cloud, ~1200 AI points (per CapCut help)",
                "Official CapCut Standard vs Pro / pricing-change pages",
            ],
            [
                "CapCut",
                "Team",
                "~$24.99 / user / month+",
                "Pro + collaboration",
                "Starts-at pricing; confirm checkout",
            ],
            [
                "Kling AI",
                "Free",
                "$0",
                "Daily credits (commonly cited: 66/day, 24h expiry)",
                "International site (kling.ai); CN site may differ",
            ],
            [
                "Kling AI",
                "Standard",
                "$10 / month (annual ~$6.60/mo equiv.; annual total often $79.20)",
                "~660 credits / month",
                "First-month promo may be lower; renew at list",
            ],
            [
                "Kling AI",
                "Pro",
                "$37 / month (annual ~$24.42/mo; annual total often $293.04)",
                "~3,000 credits / month",
                "Verify live membership page",
            ],
            [
                "Kling AI",
                "Premier",
                "$92 / month (annual ~$60.72/mo; annual total often $728.64)",
                "~8,000 credits / month",
                "Verify live membership page",
            ],
            [
                "Kling AI",
                "Ultra",
                "$180 / month (no annual option commonly reported)",
                "~26,000 credits / month",
                "Monthly-only reported as of mid-2026",
            ],
            [
                "Runway",
                "Free",
                "$0",
                "125 one-time credits",
                "Official runwayml.com/pricing",
            ],
            [
                "Runway",
                "Standard",
                "$15 / month or $12 / month billed annually",
                "625 credits / month",
                "Web app credits ≠ API credits",
            ],
            [
                "Runway",
                "Pro",
                "$35 / month or $28 / month billed annually",
                "2,250 credits / month",
                "Official pricing page",
            ],
            [
                "Runway",
                "Max",
                "$95 / month or $76 / month billed annually",
                "9,500 credits / month; unused can roll 1 month",
                "Replaces Unlimited for new subscribers",
            ],
            [
                "Runway",
                "Enterprise",
                "Custom",
                "Custom credits / SSO / support",
                "Contact sales",
            ],
        ],
    )

    add_sheet(
        wb,
        "API",
        ["service", "public_api", "billing_model", "entry_or_unit_price", "example_generation_cost", "notes"],
        [
            [
                "CapCut",
                "No public video render/generation API rate card",
                "N/A",
                "N/A",
                "N/A",
                "Consumer Pro is editor subscription, not API. Unofficial wrappers are unsupported.",
            ],
            [
                "Kling AI",
                "Yes (Open Platform / developer)",
                "Prepaid resource packages (expire); separate from membership",
                "Trial ~$9.80/100 units; production packs e.g. $700/5,000 … $7,560/60,000 units",
                "Often billed in units/second by model/resolution/audio (e.g. ~6 units/s cited for 720p silent)",
                "Membership grants no API; API units do not fund web UI. Confirm kling.dev/pricing",
            ],
            [
                "Runway",
                "Yes (dev.runwayml.com)",
                "Buy API credits; 1 credit = $0.01",
                "$0.01 / credit (sales tax may apply); commonly $10 minimum top-up",
                "gen4.5: 12 cr/s → 5s = $0.60; gen4_turbo: 5 cr/s → 5s = $0.25",
                "Official docs: docs.dev.runwayml.com/guides/pricing.md. Web plan credits unused on API.",
            ],
        ],
    )

    add_sheet(
        wb,
        "Runway_API_model_rates",
        ["model", "credits_per_second_or_unit", "usd_at_0.01", "source"],
        [
            ["gen4.5", "12 / second", "$0.12 / s", "docs.dev.runwayml.com/guides/pricing.md"],
            ["gen4_turbo", "5 / second", "$0.05 / s", "same"],
            ["aleph2", "28 / second (56 credit minimum)", "$0.28 / s", "same"],
            ["act_two", "5 / second", "$0.05 / s", "same"],
            ["seedance2 (480p/720p)", "36 / second", "$0.36 / s", "same"],
            ["seedance2 (1080p)", "40 / second", "$0.40 / s", "same"],
            ["veo3.1 (audio)", "40 / second", "$0.40 / s", "same"],
            ["veo3.1 (no audio)", "20 / second", "$0.20 / s", "same"],
            ["gen4_image", "5 @720p / 8 @1080p per image", "$0.05 / $0.08", "same"],
        ],
    )

    add_sheet(
        wb,
        "sources",
        ["service", "kind", "url", "note"],
        [
            ["CapCut", "help", "https://www.capcut.com/help/pricing-change", "Membership restructure Mar 2026"],
            [
                "CapCut",
                "guide",
                "https://capcut.com/resource/capcut-standard-vs-pro",
                "Lists Pro $19.99/mo, $179.99/yr, Team ~$24.99",
            ],
            ["Runway", "pricing", "https://runwayml.com/pricing", "Free/Standard/Pro/Max"],
            [
                "Runway",
                "credits help",
                "https://help.runwayml.com/hc/en-us/articles/15124877443219-How-do-credits-work",
                "Web vs API credit separation",
            ],
            [
                "Runway",
                "API pricing",
                "https://docs.dev.runwayml.com/guides/pricing.md",
                "1 credit = $0.01 + per-model rates",
            ],
            ["Kling", "credits policy", "https://kling.ai/docs/point-policy", "Credit acquisition rules"],
            [
                "Kling",
                "secondary (membership USD)",
                "multiple 2026 teardowns ($10/$37/$92/$180)",
                "Official membership HTML blocked in this research env (HTTP 446); re-check checkout",
            ],
            [
                "Kling",
                "secondary (API packs)",
                "developer package tables in 2026 writeups",
                "Re-check official developer pricing before spend",
            ],
        ],
    )

    add_sheet(
        wb,
        "commands",
        ["step", "command", "purpose"],
        [
            [
                "1",
                "cd /home/yoshinaga_kosuke/workspace/quiz",
                "Repo root",
            ],
            [
                "2",
                "python3 -c \"import openpyxl\" || pip install --user openpyxl",
                "Ensure openpyxl is available",
            ],
            [
                "3",
                "python3 docs/research/video-ai-pricing/build_pricing_xlsx.py",
                "Generate video-ai-pricing-compare.xlsx beside this script",
            ],
            [
                "4",
                "ls -la docs/research/video-ai-pricing/",
                "Confirm xlsx and companion files",
            ],
        ],
    )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
